from importlib.metadata import version
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from naijaledger.extract.outcome import Block, ExtractionOutcome
from naijaledger.extractions.types import DETERMINISTIC_CONFIDENCE

XLSX_METHOD_VERSION = f"openpyxl-{version('openpyxl')}"


def _cell_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_xlsx(
    data: bytes,
    *,
    content_type: str,
    content_type_conf: float,
) -> ExtractionOutcome:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    blocks: list[Block] = []
    try:
        for sheet in workbook.worksheets:
            rows: list[list[str | int | float | bool | None]] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [_cell_value(cell) for cell in row]
                if any(cell is not None and cell != "" for cell in cells):
                    rows.append(cells)
            if not rows:
                continue
            blocks.append(
                Block(
                    kind="table",
                    payload={"sheet": sheet.title, "rows": rows},
                    page=None,
                    region=None,
                )
            )
    finally:
        workbook.close()

    return ExtractionOutcome(
        method="xlsx",
        method_version=XLSX_METHOD_VERSION,
        derivation="extracted",
        confidence=DETERMINISTIC_CONFIDENCE,
        status="parsed" if blocks else "failed",
        content_type=content_type,
        content_type_conf=content_type_conf,
        blocks=blocks,
    )
